from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Initialize Excel file
excel_file =r"C:\Users\admin\Downloads\data.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Phone"])  # Header row
    wb.save(excel_file)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        
        if name and email and phone:
            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([name, email, phone])
            wb.save(excel_file)
            return "Details successfully saved!"
        else:
            return "Please fill in all fields!"

    return render_template("index.html")



if __name__ == "__main__":
    app.run(debug=True)







